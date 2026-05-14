"""Dump structure of every XLSX bible so we can refine the extraction prompt.
Outputs sheet names + first N rows of each sheet, looking for the row that
contains column headers."""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import load_workbook
from pathlib import Path

SAMPLES = Path(__file__).parent

def dump_sheet(ws, max_rows=12, max_cols=14):
    out = []
    for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
        # Skip completely empty rows
        if not any(c not in (None, "") for c in row):
            continue
        cells = []
        for c in row:
            if c is None:
                cells.append("")
            else:
                s = str(c).strip().replace("\n", " ")
                cells.append(s[:60] + ("…" if len(s) > 60 else ""))
        out.append(" | ".join(cells))
    return "\n".join(out)

xlsx_files = sorted(p for p in SAMPLES.iterdir() if p.suffix.lower() == ".xlsx" and not p.name.startswith("~$"))

for path in xlsx_files:
    print("=" * 80)
    print(f"FILE: {path.name}  ({path.stat().st_size:,} bytes)")
    print("=" * 80)
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        print(f"  ERROR: {e}")
        continue
    print(f"  Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try: dims = ws.calculate_dimension(force=True)
        except Exception: dims = "?"
        print(f"\n  --- Sheet: '{sheet_name}'  (dims: {dims}) ---")
        text = dump_sheet(ws)
        if text:
            for line in text.split("\n"):
                print(f"    {line}")
        else:
            print("    [empty]")
    wb.close()
    print()
